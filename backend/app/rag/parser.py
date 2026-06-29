"""Multi-format document parser for trip data files.

Supports:
- Excel (.xlsx): Structured trip schedules in spreadsheet format
- PDF (.pdf): Travel brochures, itinerary PDFs
- JSON (.json): Structured trip data
- Markdown (.md): Trip descriptions
"""

from typing import Any, BinaryIO
import json
import io


class DocumentParser:
    """Parse various document formats into structured trip data."""

    # Supported file extensions
    SUPPORTED_EXTENSIONS = {"xlsx", "pdf", "json", "md", "markdown"}

    @classmethod
    def validate_format(cls, filename: str) -> str:
        """Validate and return file extension."""
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext not in cls.SUPPORTED_EXTENSIONS:
            raise ValueError(
                f"不支持的文件格式: .{ext}，支持: {', '.join(cls.SUPPORTED_EXTENSIONS)}"
            )
        return ext

    @classmethod
    async def parse(cls, content: bytes, filename: str) -> dict[str, Any]:
        """Parse document content into structured data.

        Args:
            content: Raw file bytes.
            filename: Original filename (used to detect format).

        Returns:
            Dict containing parsed trip data with keys:
            - trip: dict with trip metadata
            - schedules: list of day-by-day schedules
            - raw_text: full extracted text for RAG chunking
        """
        ext = cls.validate_format(filename)

        if ext == "xlsx":
            return cls._parse_excel(content)
        elif ext == "pdf":
            return cls._parse_pdf(content)
        elif ext == "json":
            return cls._parse_json(content)
        elif ext in ("md", "markdown"):
            return cls._parse_markdown(content)
        else:
            raise ValueError(f"Unsupported format: {ext}")

    @classmethod
    def _parse_excel(cls, content: bytes) -> dict[str, Any]:
        """Parse Excel trip schedule file.

        Expected sheet structure:
        - Sheet "行程概要": overview with destination, price, etc.
        - Sheet "行程安排": day-by-day schedule with columns:
          天数 | 主题 | 类型 | 地点 | 活动 | 描述 | 餐饮 | 酒店 | 交通 | 注意事项
        """
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        trip_data = {
            "title": "",
            "destination": "",
            "duration_days": 0,
            "price_adult": None,
            "category": "国内游",
        }
        schedules = []
        all_text = []

        # Parse overview sheet if exists
        if "行程概要" in wb.sheetnames:
            ws = wb["行程概要"]
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row[0] and row[1]:
                    key, val = str(row[0]).strip(), str(row[1]).strip()
                    all_text.append(f"{key}: {val}")
                    if "目的地" in key:
                        trip_data["destination"] = val
                    elif "标题" in key or "行程名" in key:
                        trip_data["title"] = val
                    elif "天数" in key:
                        try:
                            trip_data["duration_days"] = int(val.replace("天", ""))
                        except ValueError:
                            pass
                    elif "成人价" in key:
                        try:
                            trip_data["price_adult"] = float(val)
                        except ValueError:
                            pass
                    elif "类型" in key or "分类" in key:
                        trip_data["category"] = val

        # Parse schedule sheet
        schedule_sheet_name = None
        for name in ["行程安排", "日程", "行程表", "Sheet1"]:
            if name in wb.sheetnames:
                schedule_sheet_name = name
                break

        if schedule_sheet_name:
            ws = wb[schedule_sheet_name]
            rows = list(ws.iter_rows(min_row=1, values_only=True))

            if rows:
                # First row is headers
                headers = [str(h).strip() if h else "" for h in rows[0]]

                for row in rows[1:]:
                    if not any(row):
                        continue

                    schedule = {
                        "day_number": 0,
                        "theme": "",
                        "schedule_type": "景点",
                        "location": "",
                        "activity": "",
                        "description": "",
                        "meal_included": "",
                        "hotel_name": "",
                        "transport_type": "",
                        "tips": "",
                    }

                    for i, cell in enumerate(row):
                        if i >= len(headers) or not cell:
                            continue
                        val = str(cell).strip()
                        all_text.append(val)

                        header = headers[i]
                        if "天" in header or "日" in header:
                            try:
                                schedule["day_number"] = int(
                                    val.replace("第", "").replace("天", "")
                                )
                            except ValueError:
                                pass
                        elif "主题" in header:
                            schedule["theme"] = val
                        elif "类型" in header:
                            schedule["schedule_type"] = val
                        elif "地点" in header:
                            schedule["location"] = val
                        elif "活动" in header:
                            schedule["activity"] = val
                        elif "描述" in header or "内容" in header:
                            schedule["description"] = val
                        elif "餐" in header:
                            schedule["meal_included"] = val
                        elif "酒店" in header or "住宿" in header:
                            schedule["hotel_name"] = val
                        elif "交通" in header:
                            schedule["transport_type"] = val
                        elif "注意" in header:
                            schedule["tips"] = val

                    if schedule["day_number"] > 0 or schedule["description"]:
                        schedules.append(schedule)

        return {
            "trip": trip_data,
            "schedules": schedules,
            "raw_text": "\n".join(all_text),
        }

    @classmethod
    def _parse_pdf(cls, content: bytes) -> dict[str, Any]:
        """Parse PDF travel brochure."""
        all_text = []

        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        all_text.append(text)

                    # Try to extract tables
                    tables = page.extract_tables()
                    for table in tables:
                        for row in table:
                            row_text = " | ".join(
                                str(cell) for cell in row if cell
                            )
                            all_text.append(row_text)
        except Exception:
            # Fallback: try unstructured
            try:
                from unstructured.partition.pdf import partition_pdf
                elements = partition_pdf(file=io.BytesIO(content))
                all_text = [str(el) for el in elements]
            except Exception:
                all_text = ["[PDF parsing requires pdfplumber or unstructured]"]

        return {
            "trip": {"title": "", "destination": "", "duration_days": 0},
            "schedules": [],
            "raw_text": "\n".join(all_text),
        }

    @classmethod
    def _parse_json(cls, content: bytes) -> dict[str, Any]:
        """Parse JSON structured trip data."""
        data = json.loads(content.decode("utf-8"))
        trip = data.get("trip", data.get("trip_info", {}))
        schedules = data.get("schedules", data.get("itinerary", []))

        # Ensure schedules have day_number
        for i, s in enumerate(schedules):
            if "day_number" not in s:
                s["day_number"] = i + 1

        all_text = [json.dumps(data, ensure_ascii=False, indent=2)]

        return {
            "trip": trip if isinstance(trip, dict) else {},
            "schedules": schedules,
            "raw_text": "\n".join(all_text),
        }

    @classmethod
    def _parse_markdown(cls, content: bytes) -> dict[str, Any]:
        """Parse Markdown trip description."""
        text = content.decode("utf-8")
        lines = text.split("\n")

        trip = {"title": "", "destination": "", "duration_days": 0}
        schedules = []
        current_schedule = None

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # Extract title from H1
            if line.startswith("# ") and not trip["title"]:
                trip["title"] = line[2:].strip()

            # Extract day headers
            elif line.startswith("## ") and ("天" in line or "Day" in line or "日" in line):
                if current_schedule:
                    schedules.append(current_schedule)
                current_schedule = {
                    "day_number": len(schedules) + 1,
                    "theme": line[3:].strip(),
                    "description": "",
                    "schedule_type": "景点",
                }
            elif current_schedule is not None:
                current_schedule["description"] += line + "\n"

        if current_schedule:
            schedules.append(current_schedule)

        return {
            "trip": trip,
            "schedules": schedules,
            "raw_text": text,
        }
